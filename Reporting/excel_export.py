import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import io
from typing import Dict, Any, List, Tuple

class ExcelExporter:
    """Professional Excel export service for financial statements"""
    
    # Color scheme - blue/teal theme
    COLORS = {
        'header_bg': '1F4E79',      # Dark blue
        'header_text': 'FFFFFF',     # White
        'subheader_bg': '4A90E2',    # Medium blue
        'subheader_text': 'FFFFFF',  # White
        'section_bg': 'E3F2FD',      # Light blue
        'section_text': '1565C0',    # Dark blue
        'total_bg': '81C784',        # Light green
        'total_text': '2E7D32',      # Dark green
        'border': 'BDBDBD',          # Light gray
        'alternate_bg': 'F5F5F5'     # Very light gray
    }
    
    def __init__(self):
        self.workbook = None
        self.current_row = 1
        
    def create_workbook(self, firm_name: str, start_date: str, end_date: str, granularity: str) -> openpyxl.Workbook:
        """Create a new workbook with proper styling"""
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Add metadata sheet
        self._create_metadata_sheet(firm_name, start_date, end_date, granularity)
        
        return self.workbook
    
    def _create_metadata_sheet(self, firm_name: str, start_date: str, end_date: str, granularity: str):
        """Create a metadata sheet with report information"""
        ws = self.workbook.create_sheet("Report Info")
        
        # Title
        ws['A1'] = f"{firm_name} - Financial Analysis Report"
        ws['A1'].font = Font(size=20, bold=True, color=self.COLORS['header_text'])
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header_bg'], end_color=self.COLORS['header_bg'], fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Report details
        details = [
            ("Report Period:", f"{start_date} to {end_date}"),
            ("Granularity:", granularity),
            ("Generated:", datetime.now().strftime("%B %d, %Y at %I:%M %p")),
            ("Report Type:", "Financial Statements Analysis")
        ]
        
        for i, (label, value) in enumerate(details, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True, color=self.COLORS['section_text'])
            ws[f'B{i}'].font = Font(color=self.COLORS['section_text'])
        
        # Auto-adjust column widths
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 20
    
    def add_income_statement(self, income_data: Dict[str, Dict[str, float]], sheet_name: str = "Income Statement"):
        """Add income statement with professional formatting"""
        ws = self.workbook.create_sheet(sheet_name)
        
        # Get periods and line items
        periods = list(income_data.keys())
        if not periods:
            return
            
        first_period = periods[0]
        line_items = list(income_data[first_period].keys())
        
        # Headers
        self._add_sheet_header(ws, "Income Statement", periods)
        
        # Data rows
        current_row = 4
        for line_item in line_items:
            # Line item name
            ws[f'A{current_row}'] = line_item
            ws[f'A{current_row}'].font = Font(bold=True, color=self.COLORS['section_text'])
            
            # Values for each period
            for col_idx, period in enumerate(periods, start=2):
                cell = ws[f'{get_column_letter(col_idx)}{current_row}']
                value = income_data[period].get(line_item, 0)
                cell.value = value
                
                # Format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                
                # Color coding for different line items
                if line_item == 'Revenue':
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                elif line_item == 'Expenses':
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                elif line_item == 'Net Income':
                    cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                    cell.font = Font(bold=True, color=self.COLORS['total_text'])
            
            current_row += 1
        
        # Add totals row
        self._add_totals_row(ws, current_row, periods, income_data)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, periods)
    
    def add_balance_sheet(self, balance_data: Dict[str, Dict[str, float]], sheet_name: str = "Balance Sheet"):
        """Add balance sheet with professional formatting"""
        ws = self.workbook.create_sheet(sheet_name)
        
        # Get periods and line items
        periods = list(balance_data.keys())
        if not periods:
            return
            
        first_period = periods[0]
        line_items = list(balance_data[first_period].keys())
        
        # Headers
        self._add_sheet_header(ws, "Balance Sheet", periods)
        
        # Data rows
        current_row = 4
        for line_item in line_items:
            # Line item name
            ws[f'A{current_row}'] = line_item
            ws[f'A{current_row}'].font = Font(bold=True, color=self.COLORS['section_text'])
            
            # Values for each period
            for col_idx, period in enumerate(periods, start=2):
                cell = ws[f'{get_column_letter(col_idx)}{current_row}']
                value = balance_data[period].get(line_item, 0)
                cell.value = value
                
                # Format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                
                # Color coding for different line items
                if line_item == 'Assets':
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                elif line_item == 'Liabilities':
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                elif line_item == 'Equity':
                    cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                    cell.font = Font(bold=True, color=self.COLORS['total_text'])
            
            current_row += 1
        
        # Add totals row
        self._add_totals_row(ws, current_row, periods, balance_data)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, periods)
    
    def add_cash_flow(self, cash_data: Dict[str, Dict[str, float]], sheet_name: str = "Cash Flow"):
        """Add cash flow statement with professional formatting"""
        ws = self.workbook.create_sheet(sheet_name)
        
        # Get periods and line items
        periods = list(cash_data.keys())
        if not periods:
            return
            
        first_period = periods[0]
        line_items = list(cash_data[first_period].keys())
        
        # Headers
        self._add_sheet_header(ws, "Cash Flow Statement", periods)
        
        # Data rows
        current_row = 4
        for line_item in line_items:
            # Line item name
            ws[f'A{current_row}'] = line_item
            ws[f'A{current_row}'].font = Font(bold=True, color=self.COLORS['section_text'])
            
            # Values for each period
            for col_idx, period in enumerate(periods, start=2):
                cell = ws[f'{get_column_letter(col_idx)}{current_row}']
                value = cash_data[period].get(line_item, 0)
                cell.value = value
                
                # Format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                
                # Color coding for different line items
                if line_item == 'Net Cash Change':
                    cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                    cell.font = Font(bold=True, color=self.COLORS['total_text'])
            
            current_row += 1
        
        # Add totals row
        self._add_totals_row(ws, current_row, periods, cash_data)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, periods)
    
    def _add_sheet_header(self, ws: Worksheet, title: str, periods: List[str]):
        """Add a professional header to a worksheet"""
        # Main title
        ws['A1'] = title
        ws['A1'].font = Font(size=18, bold=True, color=self.COLORS['header_text'])
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header_bg'], end_color=self.COLORS['header_bg'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(len(periods) + 1)}1')
        
        # Column headers
        ws['A2'] = "Line Item"
        ws['A2'].font = Font(bold=True, color=self.COLORS['subheader_text'])
        ws['A2'].fill = PatternFill(start_color=self.COLORS['subheader_bg'], end_color=self.COLORS['subheader_bg'], fill_type='solid')
        
        for col_idx, period in enumerate(periods, start=2):
            cell = ws[f'{get_column_letter(col_idx)}2']
            cell.value = period
            cell.font = Font(bold=True, color=self.COLORS['subheader_text'])
            cell.fill = PatternFill(start_color=self.COLORS['subheader_bg'], end_color=self.COLORS['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add borders
        self._add_borders(ws, 1, 2, len(periods) + 1)
    
    def _add_totals_row(self, ws: Worksheet, row: int, periods: List[str], data: Dict[str, Dict[str, float]]):
        """Add a totals row with proper formatting"""
        # Totals label
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True, color=self.COLORS['total_text'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['total_bg'], end_color=self.COLORS['total_bg'], fill_type='solid')
        
        # Calculate and add totals for each period
        for col_idx, period in enumerate(periods, start=2):
            cell = ws[f'{get_column_letter(col_idx)}{row}']
            total = sum(data[period].values())
            cell.value = total
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, color=self.COLORS['total_text'])
            cell.fill = PatternFill(start_color=self.COLORS['total_bg'], end_color=self.COLORS['total_bg'], fill_type='solid')
        
        # Add borders
        self._add_borders(ws, row, row, len(periods) + 1)
    
    def _add_borders(self, ws: Worksheet, start_row: int, end_row: int, end_col: int):
        """Add borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='thin', color=self.COLORS['border']),
            bottom=Side(style='thin', color=self.COLORS['border'])
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, end_col + 1):
                ws[f'{get_column_letter(col)}{row}'].border = thin_border
    
    def _adjust_column_widths(self, ws: Worksheet, periods: List[str]):
        """Auto-adjust column widths for better readability"""
        ws.column_dimensions['A'].width = 25  # Line Item column
        
        for col_idx in range(2, len(periods) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18  # Period columns
    
    def save_to_bytes(self) -> bytes:
        """Save workbook to bytes for API response"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

def create_financial_excel_report(
    firm_name: str,
    start_date: str,
    end_date: str,
    granularity: str,
    income_data: Dict[str, Dict[str, float]],
    balance_data: Dict[str, Dict[str, float]],
    cash_data: Dict[str, Dict[str, float]]
) -> bytes:
    """Create a complete financial Excel report"""
    exporter = ExcelExporter()
    exporter.create_workbook(firm_name, start_date, end_date, granularity)
    
    # Add all financial statements
    exporter.add_income_statement(income_data)
    exporter.add_balance_sheet(balance_data)
    exporter.add_cash_flow(cash_data)
    
    # Return the Excel file as bytes
    return exporter.save_to_bytes()

